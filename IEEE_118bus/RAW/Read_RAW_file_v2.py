# Read RAWM file to get bus, load, generator data
# to create a PIN excel file for ePHASORSIM
# Created by Quan Nguyen, July 5, 2018


import csv
from openpyxl import Workbook



#with open('case118_modify_genafter_FS.raw', 'r') as csvfile:
with open('case118_modify_genafter_FS_FT_Ex_1.raw', 'r') as csvfile:

    """ The following input needs to be specified by users for each RAW file """    
    # -------------------------------------------------------------------------
    bus_row                 = [4, 121]
    load_row                = [123, 275] # including PV data
    gen_row                 = [292, 345]
    shunt_row               = [277, 290]    
    # -------------------------------------------------------------------------
#    # -------------------------------------------------------------------------
#    bus_row                 = [4, 121]
#    load_row                = [123, 289] # including PV data and shunt
#    gen_row                 = [292, 345]   
#    # -------------------------------------------------------------------------





    """ Extract data from the RAW file """
    reader                  = csv.reader(csvfile, delimiter=',')
    
    """ --------- Initialize INCOMING PINS ----------- """ 
    """ Parameters """
    PLoadList               = []
    QLoadList               = []
    PpvList                 = []
    PGenList                = []       
    statusGenList           = []    
    """ Control variables """
    QpvList                 = []        
    shuntList               = []      

    """ --------- Initialize OUTGOING PINS ----------- """ 
    VmagBusList             = []
    VangBusList             = []

    
    idx                     = 0
    for row in reader:
        idx                  += 1

        """ ------------------- Extract INCOMING PINS ----------------------------- """       
        """ read LOAD and PV data """
        if  (idx >= load_row[0]) and (idx <= load_row[1]): 
           bus_number       = row[0]
           bus_number       = bus_number.replace(" ", "")
           load_ID          = row[1]
           load_ID          = load_ID.replace(" ", "")
           load_ID          = load_ID.replace("'", "")
           if (load_ID != 'PV'):
               identifier   = 'load_Z_' + str(bus_number) + '_' + load_ID
               PLoadList    += [identifier + '/P']
               QLoadList    += [identifier + '/Q']
               pre_bus      = bus_number
           elif (load_ID == 'PV'):
               if bus_number != pre_bus:
                   identifier   = 'load_Z_' + str(bus_number) + '_' + load_ID
                   PLoadList    += [identifier + '/P']
                   QLoadList    += [identifier + '/Q']

               
        """ Generator data """
        if (idx >= gen_row[0]) and (idx <= gen_row[1]):
           bus_number       = row[0]
           bus_number       = bus_number.replace(" ", "")
           gen_ID           = row[1]
           gen_ID           = gen_ID.replace(" ", "")           
           gen_ID           = gen_ID.replace("'", "")
           identifier       = 'gen_' + str(bus_number) + '_' + gen_ID
           PGenList         += [identifier + '/Tm']
           statusGenList    += [identifier + '/status']
    
        """ Switched data -> fixed shunt """
        if (idx >= shunt_row[0]) and (idx <= shunt_row[1]):
           bus_number       = row[0]
           bus_number       = bus_number.replace(" ", "")
           shunt_ID         = row[1]
           shunt_ID         = shunt_ID.replace(" ", "")           
           shunt_ID         = shunt_ID.replace("'", "")
           identifier       = 'fSH_' + str(bus_number) + '_' + shunt_ID
           shuntList        += [identifier + '/Q']  
           
    
        
        """ ------------------- Extract OUTGOING PINS ----------------------------- """   
        """ Bus Voltages """
        if (idx >= bus_row[0]) and (idx <= bus_row[1]):
           bus_number       = row[0]
           bus_number       = bus_number.replace(" ", "")
           identifier       = 'bus_' + str(bus_number)
           VmagBusList      += [identifier + '/Vmag']
           VangBusList      += [identifier + '/Vang']           
    






    
    """ Create a PIn ecel file to store all of the data extracted from the RAW file """
    filename                = 'IEEE118_PINs_1.xlsx'
    wb                      = Workbook(write_only   = True)
    ws_General              = wb.create_sheet(title = "General")
    ws_Pins                 = wb.create_sheet(title = "Pins")   
    
    
    """ --------- INCOMING PINS ----------- """ 
    """ Parameters """
    # Load row
    PLoadList               = ['incoming', 'P_Loads'] + PLoadList         
    QLoadList               = ['incoming', 'Q_Loads'] + QLoadList                               
    ws_Pins.append(PLoadList)
    ws_Pins.append(QLoadList)     
    # Pgen rows
    PGenList                = ['incoming', 'Torque'] + PGenList     
    ws_Pins.append(PGenList)
    # Pgen rows
    statusGenList           = ['incoming', 'status_Gen'] + statusGenList
    ws_Pins.append(statusGenList)

    """ Control variables """  
    # Qpv rows
    shuntList                 = ['incoming', 'shunt'] + shuntList
    ws_Pins.append(shuntList)        
    

    """ --------- OUTGOING PINS ----------- """ 
    # Bus voltage
    VmagBusList             = ['outgoing', 'Bus_voltage_mag'] + VmagBusList
    ws_Pins.append(VmagBusList)    
    VangBusList             = ['outgoing', 'Bus_voltage_ang'] + VangBusList
    ws_Pins.append(VangBusList)    

    wb.save(filename)
            
 

           
           
           
           
           
           
           
           
           
           
           
           
           
           
           
           